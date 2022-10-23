from django.conf import settings

from ..models import CsvData, Contact

import openpyxl
import csv

def print_row():
    print(f"{settings.BASE_DIR}\pochta_orders1.xlsx")

    dataframe = openpyxl.load_workbook(f"{settings.BASE_DIR}\pochta_orders1.xlsx")

    # Define variable to read sheet
    dataframe1 = dataframe.active
    
    # Iterate the loop to read the cell values
    for row in range(0, 100):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            print(col[row].value)


class Service:

    def get_data_from_csv(self):
        file: CsvData = CsvData.objects.first()

        with open(f'{file.file.path}', 'r') as f:
            csv_reader = csv.reader(f)

            for i in csv_reader:
                data = i[0].split(';')
                fio = data[1].split(' ')
                fio_len = len(fio)

                if fio_len > 1:
                    if fio_len != 3:
                        fio.append('')
                    yield [data[0], fio]

    def get_objects(self):
        data = []
        count = 0
        for i in self.get_data_from_csv():
            if count > 1000:
                Contact.objects.bulk_create(data)
                data = []
                count = 0

            data.append(Contact(phone=i[0], name=i[1][1], surname=i[1][0], patronymic=i[1][2]))

            count += 1

        Contact.objects.bulk_create(data)

    def test(self):
        with open(f'{settings.BASE_DIR}/pochta_orders1', 'r', encoding='utf-8') as f:
            lines = f.readlines(300000)
            with open(f"{settings.BASE_DIR}/new_data", 'w', encoding='UTF-8') as new:
                for line in lines:

                    new.write(line)

            for line in lines:
                print(line)


